import tkinter as tk
import openpyxl
import os
import xlutils.copy
from tkinter import filedialog
from os import getcwd
from openpyxl import load_workbook
import pandas as pd
from openpyxl.utils.exceptions import InvalidFileException
import xlrd
import numpy as np
import xlwt
import xlsxwriter
from openpyxl.utils.dataframe import dataframe_to_rows
import jinja2
import re

'''  Будет работать с двумя объектами 
  
    with style - 
    big_data_without_style - 
    
    
  '''
''' iterator that appends '''

'''  Iterate thought all cells and  save all formats in dict and then write data to excel file with applying style 
    save with no pandas, save with openpyxl 
 '''



from xlwt import easyxf

USER_CHOICE = """
Enter:
- 'o' open file to see 
- 'r' read data from file
- 'a' append data to file 
- 'w' write data to file
- 'r_range' read from specific range of data
- 'w_range' write to specific range of data
- 'show - show excel content'
Your choice: """


class pyExcel:
    def __init__(self, file_path, sheet_name, cell_range=None):
        self.file_path = file_path
        self.worksheet = self._load_workbook(file_path, sheet_name, cell_range)
        self.sheet_name = sheet_name
        self.cell_range = cell_range
        self.style = self.preserve_style()
        # self.cells_style_dict = {}

    @staticmethod
    def _open_browse_file() -> object:
        root = tk.Tk()
        root.withdraw()  # Hides the root window
        # root.wm_iconbitmap('py.ico')

        root.filename = filedialog.askopenfilename(initialdir=getcwd(),
                                                   title="Select file",
                                                   filetypes=(("excel files", "*.xlsx"),

                                                              ("all files", "*.*")))
        return root.filename

    @staticmethod
    def reformat_cell_range(cell_range):
        digit_pattern, letter_pattern = '\d+', '[a-zA-Z]+'
        start, end = cell_range.split(":")
        start_row, end_row = re.search(digit_pattern, start).group(0), \
                             re.search(digit_pattern, end).group(0)
        start_col, end_col = re.search(letter_pattern, start).group(0), \
                             re.search(letter_pattern, end).group(0)
        usecols = start_col + ":" + end_col
        return int(start_row), int(end_row), start_col, end_col, usecols

    @property
    def get_cell_range(self):
        return self.cell_range

    @property
    def get_data(self):
        return self.worksheet

    def write_data_to_cells(self, rows, columns, data):
        r_start, r_end = rows[0], rows[1]
        c_start, c_end = columns[0], columns[1]
        self.worksheet.iloc[r_start:r_end, c_start:c_end] = data

    def _load_workbook(self, file_path, sheet_name, cell_range):

        try:
            if cell_range is not None:
                worksheet = self.load_range_of_cells_excel(file_path=file_path, sheet_name=sheet_name,
                                                           cell_range=cell_range)
            else:
                worksheet = pd.read_excel(file_path)
        except FileNotFoundError:
            self.file_path = self._open_browse_file()
            worksheet = pd.read_excel(file_path, sheet_name=sheet_name)
        return worksheet

    @property
    def open_file(self):
        try:
            os.system('start EXCEL.EXE ' + self.file_path)
        except FileNotFoundError:
            return "File not found"

    @classmethod
    def load_range_of_cells_excel(cls, file_path, sheet_name, cell_range):
        start_row, end_row, start_col, end_col, usecols = cls.reformat_cell_range(cell_range)
        data = pd.read_excel(io=file_path,
                             sheet_name=sheet_name,
                             usecols=usecols,
                             nrows=end_row - start_row, skiprows=start_row - 1)
        return data

    def search_data_from_table(self, data):
        mask = self.worksheet.apply(lambda row: row.astype(str).str.contains(data).any(), axis=1)
        search_result = self.worksheet[mask]
        return search_result

    def save_to_excel(self, mode='w'):
        if mode == 'a':
            with pd.ExcelWriter(self.file_path,
                                mode=mode) as writer:
                self.to_excel(writer, sheet_name=self.sheet_name)
        else:
            rows = dataframe_to_rows(self.worksheet)
            wb = load_workbook(filename=self.file_path)
            ws = wb.active
            for r_idx, row in enumerate(rows, 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)

            wb.save(self.file_path)
            # (header_cell_style, cell_style) = self.preserve_style()
            # print(header_cell_style, cell_style)
            # props = f'font-family: "{header_cell_style["font_name"]}"; color:rgb {header_cell_style["font_color_h"]};'
            # self.worksheet = self.worksheet[]
            # print(self.file_path, self.sheet_name)
            # self.worksheet.style.set_table_styles([{'selector': 'th','props': [('background-color', 'gray')] }]).to_excel(self.file_path,index=False)
            # self.worksheet.to_excel(self.file_path, sheet_name=self.sheet_name, index=False)
            # self.set_style_of_header()

    # @staticmethod
    def set_style_of_header(self):
        from openpyxl.styles import Font
        wb = load_workbook(filename=self.file_path)
        ws = wb['эдвайзеры']
        red_font = Font(color='00FF0000', italic=True)
        for cell in ws["2:2"]:
            cell.font = red_font
        wb.save(filename=self.file_path)



    def preserve_style(self):
        'for xlrd'

        '''  for xlsx'''
        excel_version = self.file_path.split('.')[-1]
        if excel_version == 'xlsx':
            # header_cell_style, cell_style = self.__xlsx_format_handler(self.sheet_name)
            self.__xlsx_format_handler(self.sheet_name)
            # print(header_cell_style, cell_style)
        elif excel_version == 'xls':
            header_cell_style, cell_style = self.__xls_format_handler()
            print(header_cell_style, cell_style)
        # return header_cell_style, cell_style

        # return cell_style
    @staticmethod
    def get_style_cell_xlsx(cell):
        fgColor = cell.fill.fgColor.rgb
        bgColor = cell.fill.bgColor.rgb
        font_name = cell.font.name
        font_size = cell.font.size
        alignment_vertical = cell.alignment.vertical
        alignment_horizontal = cell.alignment.horizontal
        cell_style = fgColor, bgColor, font_name, font_size, alignment_vertical, alignment_horizontal
        return cell_style

    @staticmethod
    def _xls_get_style_of_cell(book, cell):
        xf = book.xf_list[cell.xf_index]
        font = book.font_list[xf.font_index]
        font_color_index = font.colour_index
        font_color = book.colour_map.get(font_color_index)
        font_name = font.name
        background_color = book.colour_map.get(xf.background.pattern_colour_index)
        text_direction = xf.alignment.text_direction
        vertical_alignment = xf.alignment.vert_align
        horizontal_alignment = xf.alignment.hor_align
        cell_style = (font_color, font_name, background_color,
                      text_direction, vertical_alignment, horizontal_alignment)
        return cell_style

    def __xlsx_format_handler(self, sheet_name):
        wb = openpyxl.load_workbook(self.file_path)
        worksheet = wb[sheet_name]
        cells_style_dict = {}
        header_cell = worksheet.cell(1, 1)

        start, end = self.cell_range.split(':')
        cell_range = worksheet[start: end]

        for row in cell_range:
            for cell in row:
                # print(cell.coordinate)
                # print(self.get_style_cell_xlsx(cell))
                fgColor, bgColor, \
                font_name, font_size, \
                alignment_vertical, alignment_horizontal = self.get_style_cell_xlsx(header_cell)
                cell_style = {"fgColor": fgColor, "bgColor": bgColor,
                              "font_name": font_name, "font_size": font_size,
                              "alignment_vertical": alignment_vertical,
                              "alignment_horizontal": alignment_horizontal}
                cells_style_dict[cell.coordinate] = cell_style
        print(cells_style_dict)


        ''' iterator that appends '''

        '''  Iterate thought all cells and  save all formats in dict and then write data to excel file with applying style 
            save with no pandas, save with openpyxl 
         '''
        # fgColor_h, bgColor_h, \
        # font_name_h, font_size_h, \
        # alignment_vertical_h, alignment_horizontal_h = self.get_style_cell_xlsx(header_cell)
        # header_cell_style = {"fgColor_h": fgColor_h, "bgColor_h": bgColor_h, "font_name_h": font_name_h, "font_size_h": font_size_h,
        # "alignment_vertical_h":alignment_vertical_h, "alignment_horizontal_h": alignment_horizontal_h}
        # cell = worksheet.cell(1, 2)


        fgColor, bgColor, \
        font_name, font_size, \
        alignment_vertical, alignment_horizontal = self.get_style_cell_xlsx(cell)
        cell_style = {"fgColor": fgColor, "bgColor": bgColor, "font_name": font_name, "font_size": font_size,
         "alignment_vertical": alignment_vertical, "alignment_horizontal": alignment_horizontal}
         # header_cell_style
        return cell_style

    def __xls_format_handler(self) -> object:
        book = xlrd.open_workbook(self.file_path, formatting_info=True)
        sheets = book.sheet_by_index(0)
        header_cell = sheets.cell(0, 0)
        font_color_header, font_name_header, background_color_header, \
        text_direction_header, vertical_alignment_header, \
        horizontal_alignment_header = self._xls_get_style_of_cell(book, header_cell)
        header_cell_style =  {"font_color_h": font_color_header, "font_name_h": font_name_header, "background_color_h":background_color_header,
        "text_direction_h": text_direction_header, "vertical_alignment_h": vertical_alignment_header,
        "horizontal_alignment_h": horizontal_alignment_header }
        cell = sheets.cell(1, 2)
        font_color, font_name, background_color, \
        text_direction, vertical_alignment, \
        horizontal_alignment = self._xls_get_style_of_cell(book, cell)
        cell_style = {"font_color": font_color, "font_name": font_name, "background_color": background_color,
        "text_direction": text_direction, "vertical_alignment": vertical_alignment,
        "horizontal_alignment": horizontal_alignment}
        return header_cell_style, cell_style





class PyExcelWithStyle:
    ''' use here openpyexcel '''
    pass

ins = pyExcel('asd.xlsx', sheet_name='эдвайзеры', cell_range='A1:A3')
wb_style = ins.preserve_style()
ins.write_data_to_cells(rows=[0, 2], columns=[0, 5], data='123')
ins.save_to_excel()
# wb_style.save('asd.xlsx')
# ins.reserve_style()
# ins.save_to_excel()
# append to file
# get_style_from_excel
# search_from_whole_dataset
# put the formula to excel file
# search from the column
# print(ins.load_range_of_cells_excel(cell_range="A5:B23"))
