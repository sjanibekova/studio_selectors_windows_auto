import xlrd
import xlwt
from openpyxl.utils.exceptions import InvalidFileException
from xlutils.copy import copy
import xlsxwriter as xlsxwriter

from helpers import pyexecelerate_to_excel
import tkinter as tk
from time import time
# from helpers import pyexecelerate_to_excel
import openpyxl
import os
from pathlib import Path, PureWindowsPath
import xlutils.copy
from tkinter import filedialog
from os import getcwd
from openpyxl import load_workbook
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
import re
import pyexcelerate as ps
import xlrd


class pyExcel:
    def __init__(self, file_path, sheet_name, cell_range=None):
        print(file_path.__str__())
        self.file_path = file_path

        self.worksheet = self._load_workbook(file_path, sheet_name, cell_range)
        self.sheet_name = sheet_name
        self.cell_range = cell_range
        self.excel_version = file_path.split('.')[-1]
        # self.cells_style_dict = {}

    @staticmethod
    def _open_browse_file() -> object:
        root = tk.Tk()
        root.withdraw()  # Hides the root window
        # root.wm_iconbitmap('py.ico')

        root.filename = filedialog.askopenfilename(initialdir=getcwd(),
                                                   title="Select file",
                                                   filetypes=(("excel files", "*.xlsx"),

                                                              ("all files", "*.*")))
        return root.filename

    @staticmethod
    def reformat_cell_range(cell_range):
        digit_pattern, letter_pattern = '\d+', '[a-zA-Z]+'
        start, end = cell_range.split(":")
        start_row, end_row = re.search(digit_pattern, start).group(0), \
                             re.search(digit_pattern, end).group(0)
        start_col, end_col = re.search(letter_pattern, start).group(0), \
                             re.search(letter_pattern, end).group(0)
        usecols = start_col + ":" + end_col
        return int(start_row), int(end_row), start_col, end_col, usecols

    @property
    def get_cell_range(self):
        return self.cell_range

    @property
    def get_data(self):
        return self.worksheet

    def write_data_to_cells(self, rows, columns, data):
        r_start, r_end = rows[0], rows[1]
        c_start, c_end = columns[0], columns[1]
        self.worksheet.iloc[r_start:r_end, c_start:c_end] = data

    def _load_workbook(self, file_path, sheet_name, cell_range):

        try:
            if cell_range is not None:
                worksheet = self.load_range_of_cells_excel(file_path=file_path, sheet_name=sheet_name,
                                                           cell_range=cell_range)
            else:
                worksheet = pd.read_excel(file_path, sheet_name=sheet_name, index_col=0)
        except FileNotFoundError:
            self.file_path = self._open_browse_file()
            worksheet = pd.read_excel(file_path, sheet_name=sheet_name, index_col=0)
        return worksheet

    @property
    def open_file(self):
        try:
            os.system('start EXCEL.EXE ' + self.file_path)
        except FileNotFoundError:
            return "File not found"

    @classmethod
    def load_range_of_cells_excel(cls, file_path, sheet_name, cell_range):
        start_row, end_row, start_col, end_col, usecols = cls.reformat_cell_range(cell_range)
        if start_row == 1:
            data = pd.read_excel(io=file_path,
                                 sheet_name=sheet_name,
                                 usecols=usecols,
                                 nrows=end_row - (start_row - 1), skiprows=start_row - 1)
        else:
            data = pd.read_excel(io=file_path,
                                 sheet_name=sheet_name,
                                 usecols=usecols,
                                 nrows=end_row - (start_row - 1), skiprows=start_row - 1, header=None)

            header = pd.read_excel(io=file_path, sheet_name=sheet_name, usecols=usecols, nrows=1)
            data.columns = header.columns
        return data

    def search_data_from_table(self, data):
        mask = self.worksheet.apply(lambda row: row.astype(str).str.contains(data).any(), axis=1)
        search_result = self.worksheet[mask]
        return search_result

    @staticmethod
    def colnum_string(n):
        string = ""
        if n == 0:
            return 'A'
        while n > 0:
            n, remainder = divmod(n - 1, 26)
            string = chr(65 + remainder) + string
        return string

    def get_address_by_value(self, data):
        coordindates = {}
        if self.excel_version == 'xls':
            wb = xlrd.open_workbook(self.file_path)
            ws = wb.sheet_by_name(sheet_name=self.sheet_name)
            for row in range(ws.nrows):
                for col in range(ws.ncols):
                    if ws.cell_value(row, col) == data:
                        coordindates[ws.cell_value(row, col)] = (self.colnum_string(col + 1) + str(row + 1))
        else:
            ''' high cost operation   '''
            start_time = time()
            wb = load_workbook(filename=self.file_path)
            ws = wb[self.sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if data == cell.value:
                        coordindates[data] = cell.coordinate

            end_time = time() - start_time
            print(end_time)
        return coordindates

    def read_with_formulas(self):
        wb = openpyxl.load_workbook(self.file_path)
        ws = wb[self.sheet_name]
        self.worksheet = pd.DataFrame(ws.values)
        print(self.worksheet)

    def delete_cell_direct_from_excel(self, coodrindates):
        pass

    def delete_column_from_dataframe(self, column_name, column_number):
        pass

    def delete_row_from_dataframe(self, row=None):
        try:
            self.worksheet = self.worksheet.reset_index().drop(row)
        except KeyError:
            print('Такой строчки не существует, индекс начинается с 0')
        finally:
            return self.worksheet

    def add_column_to_dataframe(self, position, column_name, data=None):
        self.worksheet.insert(position, column_name, data)

    @staticmethod
    def insert_row(row_num, orig_df, row_to_add):
        row_num = min(max(0, row_num), len(orig_df))
        df_part_1 = orig_df.loc[0:row_num]
        df_part_2 = orig_df.loc[row_num + 1:]
        df_final = df_part_1.append(row_to_add, ignore_index=True)
        df_final = df_final.append(df_part_2, ignore_index=True)
        return df_final


    def __add_row_to_dataframe(self, row_number, data):
        '''  Warning function !!!!!!!! '''
        if row_number > self.worksheet.index.max() + 1:
            print("Invalid row_number")
        self.worksheet = self.insert_row(row_number,self.worksheet, data)


    def add_row_to_excel(self, row_number, data=None):
        try:
            wb = load_workbook(filename=self.file_path)
            ws = wb[self.sheet_name]
            ws.insert_rows(row_number)
            if data:
                for k, v in data.items():
                    ws.cell(row=2, column=k).value = v
            wb.save(self.file_path)
        except InvalidFileException:
            print('XLS doesnt support appending a row')

    def delete_row_from_excel(self):
        ''' удалить строчку напрямую из excel файла

        for xlwx use openpyescel
        for xls - use pandas  or make converter from xlsx to xls
        '''
        # sheet.delete_rows(row[0].row, 1)

    def run_macros_in_excel(self):
        pass

    def hide_column_rows(self, rows=None, columns=None):
        pass

    def show_columns_rows(self, rows=None, columns=None):
        pass

    def save_to_excel(self, file_path, sheet_name, mode='w', style=False):
        if mode == 'a':
            with pd.ExcelWriter(file_path,
                                mode=mode) as writer:
                self.worksheet.to_excel(writer, sheet_name=sheet_name)
        else:

            if self.excel_version == 'xls':
                self.worksheet.to_excel(file_path, sheet_name=sheet_name)
            else:
                start_time = time()
                if style:
                    ''' openpyxl  '''
                    self.worksheet.rename(columns=lambda x: re.sub('Unnamed: \d', '', x), inplace=True)
                    rows = dataframe_to_rows(self.worksheet)
                    print(self.worksheet)
                    wb = load_workbook(filename=self.file_path)
                    ws = wb[self.sheet_name]
                    for r_idx, row in enumerate(rows, 1):
                        for c_idx, value in enumerate(row, 1):
                            try:
                                ws.cell(row=r_idx, column=c_idx, value=value)
                            except AttributeError:
                                print('Cannot preserve style ')
                                pyexecelerate_to_excel(file_path, self.worksheet, sheet_name)
                    wb.save(self.file_path)


                else:
                    pyexecelerate_to_excel(file_path, self.worksheet, sheet_name)

                end_time = time() - start_time
                print(end_time)


'''  конвертер xls - xlsx  
      
'''
''' 1) Чтение данных '''
# large data
# ins = pyExcel('test_file.xlsx', sheet_name='Sheet 1')

# small data
# file_path = r'C:\SeikoLab\studio_selectors_windows_auto\src\snipping_tool\asdq.xlsx'
# ins = pyExcel(file_path, sheet_name='эдвайзеры')
# print(ins.get_data)


# xls data

# ins = pyExcel('sdf123213.xls', sheet_name='112354546545')
# print(ins.get_data)

''' 2) Запись данных в датафрейм'''
# ins.write_data_to_cells(rows=[1, 3], columns=[0, 3], data='-----------------------')

''' 3) Чтение определенного диапазона данных (exmpl: A1:A3) '''
# ins = pyExcel('asdq.xlsx', sheet_name='эдвайзеры', cell_range='A1:C3')


''' 4) Запись в определенный диапазон данных (Example: A1:A3)'''
# !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!

''' 5) Сохранение стилей при изменении контента файла '''
# ins.save_to_excel('sdf123213.xls', sheet_name='112354546545', style=True)
''' 6) Умение работать с office 365 ????? (не нашла исходников)'' 
(не реализовано)

# 7) Умение работать с данными (100 000 – 200 000 данных без сильных задержек)
# небольшая оптимизация pyecelerate'''

''' 8) Понимание формулы в excel файле при чтении (не реализовано еще) '''
# print(ins.read_with_formulas()) (исправить индексацию если нужно)

''' 9) Определение объединенных ячеек , а что с ними дальше делать непонятно)   '''
# ins = pyExcel(r'C:\SeikoLab\studio_selectors_windows_auto\src\snipping_tool\asdq.xlsx', sheet_name='эдвайзеры', cell_range='A1:B2')
# print(ins.get_data)
# ins.write_data_to_cells(rows=[1, 3], columns=[0, 3], data='-----------sdfsdfsdf------------')
# ins.save_to_excel('asdq.xlsx', sheet_name='эдвайзеры', style=True)
# read whole data


''' удалить строку  '''
# ins = pyExcel(r'C:\SeikoLab\studio_selectors_windows_auto\src\snipping_tool\asdq.xlsx', sheet_name='эдвайзеры', cell_range='A3:C3')
# print(ins.get_data)
# a = ins.delete_row_from_dataframe(row=1)
# print(ins.get_data)
# ins.save_to_excel(r'C:\SeikoLab\studio_selectors_windows_auto\src\snipping_tool\asdq.xlsx', sheet_name='эдвайзеры')

''' функции добавить колонку с данными'''
# ins = pyExcel(r'C:\SeikoLab\studio_selectors_windows_auto\src\snipping_tool\asdq.xlsx', sheet_name='эдвайзеры',
#               cell_range='A1:C2')
# ins.add_column_to_dataframe(3,'123', None)
# print(ins.get_data)
# ins.save_to_excel(r'C:\SeikoLab\studio_selectors_windows_auto\src\snipping_tool\asdq.xlsx', sheet_name='эдвайзеры')



''' функции добавить строку в датафрейм  WARNIING!!!!!!!!!!!!!
Затирает последнюю строчку 

'''
# ins = pyExcel(r'C:\SeikoLab\studio_selectors_windows_auto\src\snipping_tool\asdq.xlsx', sheet_name='эдвайзеры',
#               cell_range='A1:C2')
# print(ins.get_data)
# data = {2:'asdasdasdasdasd'}
# ''' starts from 0 (0-это название колонки) '''
# ins.__add_row_to_dataframe(row_number=0, data=data)
# print(ins.get_data)
# ins.save_to_excel(r'C:\SeikoLab\studio_selectors_windows_auto\src\snipping_tool\asdq.xlsx', sheet_name='эдвайзеры')


# print(ins.get_data)
# ins.save_to_excel('test_file.xlsx', sheet_name='Sheet 1')
# ins.get_address_by_value(-1.20700056906294)

# 

# ins.write_data_to_cells(rows=[1, 3], columns=[0, 3], data='-----------------------')
# print(ins.get_data)
# ins.save_to_excel()

# print(ins.get_address_by_value(-0.0646207837037598))

# Добавить строку в эксель
''' add_row_to_excel'''
# ins = pyExcel(r'C:\SeikoLab\studio_selectors_windows_auto\src\snipping_tool\asdq.xlsx', sheet_name='эдвайзеры',cell_range='A1:C2')
# print(ins.get_data)
# ins.add_row_to_excel(2, {1: "asdasd", 2:'asdasd'})


''' add_row_to_excel  xls'''
# ins = pyExcel(r'C:\SeikoLab\studio_selectors_windows_auto\src\snipping_tool\asd.xls', sheet_name='1',cell_range='A1:C2')
# print(ins.get_data)
# ins.add_row_to_excel(2, {1: "asdasd", 2:'asdasd'})


# https://www.excelpython.org/
''' удалить строку direct from excel '''
ins = pyExcel(r'C:\SeikoLab\studio_selectors_windows_auto\src\snipping_tool\asdq.xlsx', sheet_name='эдвайзеры', cell_range='A3:C3')
print(ins.get_data)
a = ins.delete_row_from_excel(row=1)
print(ins.get_data)
ins.save_to_excel(r'C:\SeikoLab\studio_selectors_windows_auto\src\snipping_tool\asdq.xlsx', sheet_name='эдвайзеры')

