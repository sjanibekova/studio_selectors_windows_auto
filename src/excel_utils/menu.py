import tkinter as tk
import openpyxl
import os
from tkinter import filedialog
from os import getcwd
from openpyxl import load_workbook
import pandas as pd
USER_CHOICE = """
Enter:
- 'o' open file to see 
- 'r' read data from file
- 'a' append data to file 
- 'w' write data to file
- 'r_range' read from specific range of data
- 'w_range' write to specific range of data
- 'show - show excel content'
Your choice: """


def menu():
    file_path = 'D:\Руководители_2_и_3_курс.xlsx'
    # excel.create_excel_file()
    user_input = input(USER_CHOICE)
    wb = load_workbook('D:\Руководители_2_и_3_курс.xlsx')
    while user_input != 'q':
        if user_input == 'o':
           wb =  load_file('D:\Руководители_2_и_3_курс.xlsx')
           break
        if user_input == 'r':
           wb = load_workbook('D:\Руководители_2_и_3_курс.xlsx')
           show_data_from_excel(file_name='D:\Руководители_2_и_3_курс.xlsx', wb=wb)
           break
        elif user_input == 'w':
            ws = choose_sheet(wb)
            range = input("Enter the range that you want read: ex: A1:C3")
            write_to_specific_range_of_cells(wb, ws, range)

        elif user_input == 'r_range':
            ws = choose_sheet(wb)
            range = input("Enter the range that you want read: ex: A1:C3")
            get_specific_range_of_data(ws, range)

        # elif user_input == 'r_range':
        #     get_range_of_cells()
        #     read_from_excel()
        # elif user_input == 'w_range':
        #     get_range_of_cells()
        #     write_to_excel()
        # user_input = input(USER_CHOICE)
        # if user_input == 'q':
        #     break


def load_file(file_path):
    try:
        wb = load_workbook(file_path)
        os.system('start EXCEL.EXE '+ file_path)
    except FileNotFoundError:
        root = tk.Tk()
        root.withdraw()  # Hides the root window
        # root.wm_iconbitmap('py.ico')

        root.filename = filedialog.askopenfilename(initialdir=getcwd(),
                                                     title="Select file",
                                                     filetypes=(("excel files", "*.xlsx"),
                                                                ("all files", "*.*")))

        wb = load_workbook(root.filename)
    return wb





def show_data_from_excel(file_name=None,wb=None):
    if len(wb.worksheets) > 1:
        worksheets = get_sheet_names(wb)
        print(worksheets)
        sheet_name = input('Enter sheet name ')
        if sheet_name in worksheets:
            data = pd.read_excel(file_name, engine='openpyxl', sheet_name=sheet_name)
            print(data.head())
    else:
        print('Not found')




def choose_sheet(wb):
    worksheets = get_sheet_names(wb)
    print(worksheets)
    sheet_name = input('Enter sheet name ')
    if sheet_name in worksheets:
        ws = wb[sheet_name]
        return ws
    else:
        print('No sheet with that name')


def get_sheet_names(wb):
    print(wb)
    worksheets = list(map(lambda x: str(x).split(' ')[1].replace('>', '').replace('"', ""), wb.worksheets))
    return worksheets
# def write_to_file(wb):




def write_to_specific_range_of_cells(wb, wbsheet, cell_range):
    input_data = input('type what you want to input ')
    start, end = cell_range.split(':')
    cell_range = wbsheet[start : end]
    for row in cell_range:
        for cell in row:
            cell.value = input_data
    save_wb(wb, 'asd.xlsx')



def get_specific_range_of_data(wbsheet, cell_range):
    start, end = cell_range.split(':')
    cell_range = wbsheet[start : end]
    print(len(cell_range))
    if len(cell_range) > 1:
        for row in cell_range:
            for cell in row:
                print(cell.value)


def create_sheets(wb, sheet_name_list):
    # Adds the sheets in the sheet_name_list to the workbook
    for sheet_name in sheet_name_list:
        wb.create_sheet(sheet_name)



def save_wb(wb, filename):
    # Save a workbook
    wb.save(filename)
menu()



from time import time
# test execution time


import tkinter as tk
import openpyxl
import os
from tkinter import filedialog
from os import getcwd
from openpyxl import load_workbook
import pandas as pd
from openpyxl.utils.exceptions import InvalidFileException

USER_CHOICE = """
Enter:
- 'o' open file to see 
- 'r' read data from file
- 'a' append data to file 
- 'w' write data to file
- 'r_range' read from specific range of data
- 'w_range' write to specific range of data
- 'show - show excel content'
Your choice: """


class pyExcel:
    def __init__(self, file_path, sheet_name):
        self.file_path = file_path
        self.wb = load_workbook(file_path)
        self.sheet_name = sheet_name
        self.worksheet = self.wb[sheet_name]

    @staticmethod
    def _open_browse_file():
        root = tk.Tk()
        root.withdraw()  # Hides the root window
        # root.wm_iconbitmap('py.ico')

        root.filename = filedialog.askopenfilename(initialdir=getcwd(),
                                                   title="Select file",
                                                   filetypes=(("excel files", "*.xlsx"),

                                                              ("all files", "*.*")))
        return root.filename

    @classmethod
    def load_workbook(cls):

        try:
            wb = load_workbook(cls.file_path)
        except FileNotFoundError:
            cls.file_path = cls._open_browse_file()
            wb = load_workbook(cls.file_path)
            print(wb)
        return wb

    def open_file(self):
        try:
            os.system('start EXCEL.EXE ' + self.file_path)
        except FileNotFoundError:
            return "File not found"

    def read_data_from_excel_sheet(self):
        if len(self.wb.worksheets) > 1:
            worksheets = self.get_sheet_names
            print(worksheets)
            sheet_name = self.sheet_name
            if sheet_name in worksheets:
                data = pd.DataFrame(self.worksheet.values)
                print(data)

    @property
    def get_sheet_names(self):
        print(self.wb)
        worksheets = list(map(lambda x: str(x).split(' ')[1].replace('>', '').replace('"', ""), self.wb.worksheets))
        return worksheets


ins = pyExcel('dataset.xls', '1')
print(ins.read_data_from_excel_sheet())
